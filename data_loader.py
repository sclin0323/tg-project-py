"""
data_loader.py — 從 Excel 載入並計算 PPT 需要的所有衍生資料 (v2: 適配新版 WBS schema)

新版 WBS schema:
  Sheet name: 'WBS清單'
  11 欄: WBS編號, 任務類型, 工作項目, 負責人, 預計開始日, 預計結束日,
         實際開始日, 實際完成日, 前置任務, 狀態, 備註
  三階編號: '1' / '1.1' / '1.1.1'
  L1 / L2 為群組列, 只填 WBS編號 + 任務類型 + 工作項目, 其他欄位空白
  L3 為實際任務, 完整欄位
  里程碑: 工作項目以 ★ 開頭的 L3 任務

回傳的結構 (report):
{
  "report_date":  date,
  "next_report":  date,
  "headline":     str,
  "overall":      "RED|AMBER|GREEN",
  "stats":  {...},
  "projects": [project_obj, ...],
  "decisions": [decision, ...],
}
"""
from __future__ import annotations
import os
from datetime import datetime, timedelta, date
import openpyxl
import pandas as pd

# ── 設定 ─────────────────────────────────────────────
LOOKBACK_DAYS  = 14
LOOKAHEAD_DAYS = 14

STATUS_DONE    = "完成"
STATUS_ACTIVE  = ("進行中",)
STATUS_DELAYED = "延遲"
STATUS_RISK_S  = "風險"
STATUS_RISK    = (STATUS_DELAYED, STATUS_RISK_S)
STATUS_SKIP    = ("取消", "暫緩", "暫緩執行")
STATUS_WAIT    = "待開始"

HEALTH_LABELS = {"RED": "警報", "AMBER": "注意", "GREEN": "健康"}


# ──────────────────────────────────────────────────────
# 基礎工具
# ──────────────────────────────────────────────────────
def _parse(v):
    """把任意值轉成 date 或 None"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        v = v.strip()
        if not v or v in ("NaT", "nan", "None"):
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


def _str(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def _level(wbs_no: str) -> int:
    """WBS 編號的層級 (1/1.1/1.1.1 → 1/2/3)"""
    if not wbs_no:
        return 0
    return wbs_no.count('.') + 1


def _is_milestone(name: str) -> bool:
    """工作項目以 ★ 開頭 → 里程碑"""
    return name.startswith("★") if name else False


# ──────────────────────────────────────────────────────
# 載入資料
# ──────────────────────────────────────────────────────
def load_master(excel_dir: str) -> list[dict]:
    """讀總表 WBS清單總表"""
    path = os.path.join(excel_dir, "專案管理總表SDC.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["WBS清單總表"]
    out = []
    for r in range(2, ws.max_row + 1):
        name = _str(ws.cell(row=r, column=2).value)
        if not name:
            continue
        out.append({
            "start":    _parse(ws.cell(row=r, column=1).value),
            "name":     name,
            "stage":    _str(ws.cell(row=r, column=3).value),
            "owner":    _str(ws.cell(row=r, column=4).value),
            "target":   _parse(ws.cell(row=r, column=5).value),
            "status":   _str(ws.cell(row=r, column=6).value),
            "wbs_file": _str(ws.cell(row=r, column=7).value),
        })
    return out


def load_wbs(excel_dir: str, wbs_file: str) -> list[dict]:
    """讀單個 WBS Excel；總表 G 欄是 'X.xlsx'，實體可能是 'WBS_X.xlsx'"""
    candidates = [wbs_file, f"WBS_{wbs_file}"]
    src = None
    for c in candidates:
        p = os.path.join(excel_dir, c)
        if os.path.exists(p):
            src = p
            break
    if not src:
        raise FileNotFoundError(f"找不到 WBS 檔案: {wbs_file}")

    wb = openpyxl.load_workbook(src, data_only=True)
    # 找 WBS 工作表 (兼容 'WBS清單' 和舊版 'WBS')
    sheet_name = None
    for sn in ("WBS清單", "WBS"):
        if sn in wb.sheetnames:
            sheet_name = sn
            break
    if not sheet_name:
        raise ValueError(f"{src} 找不到 'WBS清單' 或 'WBS' 工作表")

    ws = wb[sheet_name]
    tasks = []
    for r in range(2, ws.max_row + 1):
        wbs_no = _str(ws.cell(row=r, column=1).value)
        if not wbs_no:
            continue
        t_type = _str(ws.cell(row=r, column=2).value)
        name   = _str(ws.cell(row=r, column=3).value)
        if not name:
            continue

        lvl = _level(wbs_no)
        # L1 / L2 是群組,不會有日期、負責人;L3 才是實際任務
        is_group = (lvl in (1, 2))

        tasks.append({
            "wbs_no":      wbs_no,
            "level":       lvl,
            "task_type":   t_type,
            "name":        name,
            "owner":       _str(ws.cell(row=r, column=4).value),
            "plan_start":  _parse(ws.cell(row=r, column=5).value),
            "plan_end":    _parse(ws.cell(row=r, column=6).value),
            "actual_start": _parse(ws.cell(row=r, column=7).value),
            "actual_end":  _parse(ws.cell(row=r, column=8).value),
            "pred":        _str(ws.cell(row=r, column=9).value),
            "status":      _str(ws.cell(row=r, column=10).value),
            "note":        _str(ws.cell(row=r, column=11).value),
            "is_group":    is_group,
            "is_milestone": _is_milestone(name),
            # 為了相容舊 builder 的欄位名 (start/end/actual)
            "start":       _parse(ws.cell(row=r, column=5).value),
            "end":         _parse(ws.cell(row=r, column=6).value),
            "actual":      _parse(ws.cell(row=r, column=8).value),
        })
    return tasks


# ──────────────────────────────────────────────────────
# 健康度
# ──────────────────────────────────────────────────────
def compute_health(start: date, target: date, tasks: list[dict],
                   today: date) -> dict:
    real = [t for t in tasks
            if not t["is_group"] and t["status"] not in STATUS_SKIP]
    if not real or not start or not target:
        return {
            "level": "GREEN", "label": "健康",
            "done_pct": 0, "elapsed_pct": 0, "drift_pct": 0,
            "forecast_end": target, "forecast_drift": 0,
            "forecast_available": False,
        }

    total = len(real)
    done  = sum(1 for t in real if t["status"] == STATUS_DONE)
    risk  = sum(1 for t in real if t["status"] in STATUS_RISK)
    delayed = sum(1 for t in real if t["status"] == STATUS_DELAYED)

    total_days = max(1, (target - start).days)
    elapsed    = max(0, min(total_days, (today - start).days))
    elapsed_pct = elapsed / total_days
    done_pct    = done / total
    drift       = done_pct - elapsed_pct

    if   drift >= -0.05: level = "GREEN"
    elif drift >= -0.15: level = "AMBER"
    else:                level = "RED"

    # 風險加成:
    #   有任何「延遲」狀態 → AMBER 以上
    #   ≥3 個延遲          → RED
    #   ≥2 個風險/延遲     → AMBER
    if delayed >= 3:
        level = "RED"
    elif delayed >= 1 and level == "GREEN":
        level = "AMBER"
    elif risk >= 2 and level == "GREEN":
        level = "AMBER"
    elif risk >= 4 and level == "AMBER":
        level = "RED"

    # 預估完工
    forecast_end = target
    forecast_drift = 0
    forecast_available = False
    if done > 0 and elapsed >= 7 and done_pct >= 0.10:
        velocity = elapsed / done
        remaining = (total - done) * velocity
        forecast_end = today + timedelta(days=int(remaining))
        forecast_drift = (forecast_end - target).days
        cap = 60
        if forecast_drift > cap:
            forecast_drift, forecast_end = cap, target + timedelta(days=cap)
        elif forecast_drift < -cap:
            forecast_drift, forecast_end = -cap, target - timedelta(days=cap)
        forecast_available = True

    return {
        "level":              level,
        "label":              HEALTH_LABELS[level],
        "done_pct":           round(done_pct * 100, 1),
        "elapsed_pct":        round(elapsed_pct * 100, 1),
        "drift_pct":          round(drift * 100, 1),
        "forecast_end":       forecast_end,
        "forecast_drift":     forecast_drift,
        "forecast_available": forecast_available,
    }


# ──────────────────────────────────────────────────────
# 里程碑 — 以第 1 層 WBS (1, 2, 3...) 為「階段」
# 每個階段聚合底下所有第 3 層任務,得出時間區間與進度
# ──────────────────────────────────────────────────────
def _phase_root(wbs_no: str) -> str:
    """取得 WBS 編號的第一層根, 例如 '1.2.3' → '1', '4.5' → '4'"""
    if not wbs_no:
        return ""
    return wbs_no.split(".")[0]


def extract_milestones(tasks: list[dict]) -> list[dict]:
    """把第 1 層 (1, 2, 3...) 當成一個「里程碑/階段」,
       時間區間 = 旗下所有 L3 任務的 min(plan_start) ~ max(plan_end),
       進度 = 加權平均 (完成 1.0, 風險/延遲 0.5, 其他 0)。

       階段狀態判定:
         - 全部完成      → 完成
         - 有延遲        → 延遲
         - 有風險        → 風險
         - 有進行中或部分完成 → 進行中
         - 否則          → 待開始
    """
    # ── Step 1: 拿到所有「第 1 層群組列」當骨架 (才能保留原始順序與名稱) ──
    phase_rows = {t["wbs_no"]: t for t in tasks if t.get("level") == 1}

    # ── Step 2: 依第一層根 (1, 2, 3...) 把所有 L3 任務分群 ──
    phase_tasks: dict[str, list] = {}
    for t in tasks:
        # 只看真正會有日期的 L3 任務, 並排除取消/暫緩
        if t.get("level") != 3:
            continue
        if t["status"] in STATUS_SKIP:
            continue
        root = _phase_root(t["wbs_no"])
        if not root:
            continue
        phase_tasks.setdefault(root, []).append(t)

    # ── Step 3: 聚合 ──
    out = []
    # 依 WBS 編號數字順序輸出 (保證 1, 2, 3, ... 10 而不是 1, 10, 2)
    for root in sorted(phase_rows.keys(), key=lambda x: int(x) if x.isdigit() else 99):
        group_row = phase_rows[root]
        children = phase_tasks.get(root, [])

        # 沒有 L3 子任務的階段跳過 (例如只有 L1 沒有 L2/L3)
        if not children:
            continue

        # 名稱: 用第 1 層的工作項目, 並清掉可能殘留的 "★" 或 "(里程碑)" 字樣
        name = group_row["name"].replace("★", "").strip()
        name = name.replace("(里程碑)", "").replace("階段", "").strip()
        if not name:
            name = f"階段 {root}"

        # 時間區間
        starts = [t["plan_start"] for t in children if t["plan_start"]]
        ends   = [t["plan_end"]   for t in children if t["plan_end"]]
        if not starts or not ends:
            continue
        phase_start = min(starts)
        phase_end   = max(ends)

        # 進度: 加權平均 (完成 1.0, 風險/延遲 0.5, 其他 0)
        total = len(children)
        score = 0.0
        n_done = n_active = n_delayed = n_risk = n_wait = 0
        for t in children:
            st = t["status"]
            if st == STATUS_DONE:
                score += 1.0;  n_done += 1
            elif st == STATUS_DELAYED:
                score += 0.5;  n_delayed += 1
            elif st == STATUS_RISK_S:
                score += 0.5;  n_risk += 1
            elif st in STATUS_ACTIVE:
                score += 0.5;  n_active += 1
            else:
                n_wait += 1
        progress = int(round(score / total * 100)) if total else 0

        # 階段狀態 (依優先順序判定)
        if n_done == total:
            status = "完成"
        elif n_delayed > 0:
            status = "延遲"
        elif n_risk > 0:
            status = "風險"
        elif n_active > 0 or n_done > 0:
            status = "進行中"
        else:
            status = "待開始"

        # 負責人: 取階段內出現次數最多的負責人 (排除空白)
        owner_counts: dict[str, int] = {}
        for t in children:
            ow = t["owner"]
            if ow:
                owner_counts[ow] = owner_counts.get(ow, 0) + 1
        owner = max(owner_counts, key=owner_counts.get) if owner_counts else ""

        out.append({
            "wbs_no":  root,
            "name":    name,
            "start":   phase_start,
            "end":     phase_end,
            "status":  status,
            "progress": progress,
            "owner":   owner,
            "total":   total,
            "done":    n_done,
        })
    return out


# ──────────────────────────────────────────────────────
# 本期動態 (給第二頁四區塊用)
# ──────────────────────────────────────────────────────
def collect_activity(tasks: list[dict], today: date) -> dict:
    """收集第二頁要顯示的四個任務分類:
       delayed   — 狀態 = 延遲
       risk      — 狀態 = 風險
       active    — 狀態 = 進行中
       recent    — 近 N 天實際完成的任務
    """
    lookback = today - timedelta(days=LOOKBACK_DAYS)

    delayed, risk, active, recent = [], [], [], []

    for t in tasks:
        if t["is_group"]:
            continue
        if t["status"] in STATUS_SKIP:
            continue
        st = t["status"]

        # 延遲 / 風險 一律顯示 (不論任務類型)
        if st == STATUS_DELAYED:
            delayed.append(t)
            continue
        if st == STATUS_RISK_S:
            risk.append(t)
            continue

        # 「專案管理」類的全期任務 (例如「每週例會」、「風險登錄簿維護」)
        # 不放入「進行中」清單,避免讓管理活動稀釋了實際工作項目
        is_admin = t.get("task_type") == "專案管理"

        if st in STATUS_ACTIVE and not t["is_milestone"] and not is_admin:
            active.append(t)
        elif (st == STATUS_DONE and t["actual_end"]
              and lookback <= t["actual_end"] <= today
              and not t["is_milestone"] and not is_admin):
            recent.append(t)

    # 排序
    # 延遲: 逾期最久的優先
    def _days_late(t):
        if t["plan_end"] and t["plan_end"] < today:
            return (today - t["plan_end"]).days
        return 0
    delayed.sort(key=lambda x: -_days_late(x))
    # 風險: 計畫結束日近的優先
    risk.sort(key=lambda x: x["plan_end"] or date.max)
    # 進行中: 計畫結束日近的優先
    active.sort(key=lambda x: x["plan_end"] or date.max)
    # 近期完成: 完成日最新的優先
    recent.sort(key=lambda x: x["actual_end"], reverse=True)

    return {
        "delayed": delayed,
        "risk":    risk,
        "active":  active,
        "recent":  recent,
    }


# ──────────────────────────────────────────────────────
# 裁示事項
# ──────────────────────────────────────────────────────
def derive_decisions(project_name: str, tasks: list[dict],
                     today: date) -> list[dict]:
    out = []
    for t in tasks:
        if t["is_group"] or t["status"] not in STATUS_RISK:
            continue
        end = t["plan_end"]
        days_late = (today - end).days if (end and not t["actual_end"]) else 0

        # 嚴重度判斷
        if t["status"] == STATUS_DELAYED and days_late > 21:
            sev = "高"
        elif t["status"] == STATUS_DELAYED or days_late > 7:
            sev = "中"
        elif t["status"] == STATUS_RISK_S:
            sev = "中"
        else:
            sev = "低"

        # 議題敘述 — 包含 WBS 編號方便定位
        if days_late > 0:
            issue = f"{t['wbs_no']} {t['name']}（已逾期 {days_late} 天）"
        else:
            issue = f"{t['wbs_no']} {t['name']}（狀態:{t['status']}）"

        if sev == "高":
            ask = "需裁示：加派人力 / 縮減範圍 / 調整上線時程"
        elif sev == "中":
            ask = "需協助：釐清阻礙因素,評估是否調整計畫"
        else:
            ask = "持續追蹤進度"

        out.append({
            "severity":  sev,
            "project":   project_name,
            "wbs_no":    t["wbs_no"],
            "name":      t["name"],
            "issue":     issue,
            "impact":    t["note"] or "可能影響後續任務啟動",
            "owner":     t["owner"] or "未指派",
            "ask":       ask,
            "days_late": days_late,
        })

    sev_order = {"高": 0, "中": 1, "低": 2}
    out.sort(key=lambda d: (sev_order[d["severity"]], -d["days_late"]))
    return out


# ──────────────────────────────────────────────────────
# 一句話總結
# ──────────────────────────────────────────────────────
def make_summary(p: dict) -> str:
    h = p["health"]
    n_delayed = p["delayed"]
    n_risk    = p["risk_count"]
    drift = h["drift_pct"]

    if h["level"] == "RED":
        if n_delayed >= 3:
            return f"{n_delayed} 個任務已延遲,影響上線時程,需立即裁示"
        if n_delayed > 0:
            return f"進度落後 {abs(round(drift))}%,{n_delayed} 個任務已延遲需處理"
        return f"進度嚴重落後 {abs(round(drift))}%,需採取補救措施"
    if h["level"] == "AMBER":
        if n_delayed > 0:
            return f"{n_delayed} 個任務延遲、{n_risk} 個風險,整體仍可控"
        if n_risk > 0:
            return f"進度大致符合,{n_risk} 個任務出現風險訊號,需持續關注"
        return f"進度略落後 {abs(round(drift))}%,仍可追回"
    if h["done_pct"] > h["elapsed_pct"] + 5:
        return f"進度超前約 {round(h['done_pct'] - h['elapsed_pct'])}%,依計畫順利推進"
    return "進度依計畫推進,整體狀況良好"


# ──────────────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────────────
def build_report(excel_dir: str, today: date | None = None) -> dict:
    today = today or date.today()
    master = load_master(excel_dir)

    projects = []
    all_decisions = []

    for m in master:
        try:
            tasks = load_wbs(excel_dir, m["wbs_file"])
        except FileNotFoundError as e:
            print(f"  ⚠  {e}")
            continue

        # 自動偵測逾期 (預定結束日 < 今天 且 未完成 且 不是 延遲/風險/取消/暫緩)
        for t in tasks:
            if t["is_group"]:
                continue
            if t["status"] in (STATUS_DONE, *STATUS_RISK, *STATUS_SKIP):
                continue
            if t["plan_end"] and t["plan_end"] < today and not t["actual_end"]:
                t["status"] = STATUS_DELAYED

        real = [t for t in tasks if not t["is_group"] and t["status"] not in STATUS_SKIP]
        total = len(real)
        done    = sum(1 for t in real if t["status"] == STATUS_DONE)
        inp     = sum(1 for t in real if t["status"] in STATUS_ACTIVE)
        delayed = sum(1 for t in real if t["status"] == STATUS_DELAYED)
        risk_n  = sum(1 for t in real if t["status"] == STATUS_RISK_S)
        wait    = sum(1 for t in real if t["status"] == STATUS_WAIT)

        health = compute_health(m["start"], m["target"], tasks, today)
        milestones = extract_milestones(tasks)
        activity = collect_activity(tasks, today)
        decisions = derive_decisions(m["name"], tasks, today)
        all_decisions.extend(decisions)

        p = {
            "name":        m["name"],
            "owner":       m["owner"],
            "stage":       m["stage"],
            "status":      m["status"],
            "start":       m["start"],
            "target":      m["target"],
            "total":       total,
            "done":        done,
            "in_progress": inp,
            "delayed":     delayed,
            "risk_count":  risk_n,
            "risk":        delayed + risk_n,    # 為相容舊欄位 (延遲+風險 合計)
            "not_started": wait,
            "health":      health,
            "milestones":  milestones,
            "activity":    activity,
            "decisions":   decisions,
            "tasks":       tasks,
        }
        p["summary"] = make_summary(p)
        projects.append(p)

    sev_order = {"高": 0, "中": 1, "低": 2}
    all_decisions.sort(key=lambda d: (sev_order[d["severity"]], -d["days_late"]))

    red   = sum(1 for p in projects if p["health"]["level"] == "RED")
    amber = sum(1 for p in projects if p["health"]["level"] == "AMBER")
    green = sum(1 for p in projects if p["health"]["level"] == "GREEN")

    if red > 0:
        headline = f"{red} 個專案需立即關注"
        overall = "RED"
    elif amber > 0:
        headline = f"{amber} 個專案需注意觀察"
        overall = "AMBER"
    else:
        headline = "全部專案進度健康"
        overall = "GREEN"

    return {
        "report_date":  today,
        "next_report":  today + timedelta(days=7),
        "headline":     headline,
        "overall":      overall,
        "stats": {
            "total_projects":    len(projects),
            "red_projects":      red,
            "amber_projects":    amber,
            "green_projects":    green,
            "total_tasks":       sum(p["total"]       for p in projects),
            "done_tasks":        sum(p["done"]        for p in projects),
            "in_progress_tasks": sum(p["in_progress"] for p in projects),
            "delayed_tasks":     sum(p["delayed"]     for p in projects),
            "risk_tasks":        sum(p["risk"]        for p in projects),
            "decision_count":    sum(1 for d in all_decisions if d["severity"] in ("高","中")),
        },
        "projects":  projects,
        "decisions": all_decisions,
    }


# ── 自我驗證 ──────────────────────────────────────────
if __name__ == "__main__":
    import sys
    excel_dir = sys.argv[1] if len(sys.argv) > 1 else "./excel"
    r = build_report(excel_dir, today=date(2026, 5, 25))
    print(f"\n📊 整體：{r['overall']} | {r['headline']}")
    print(f"   報告日:{r['report_date']} | 下次:{r['next_report']}")
    print(f"   紅 {r['stats']['red_projects']} / 黃 {r['stats']['amber_projects']} / 綠 {r['stats']['green_projects']}")
    print(f"\n📋 各專案：")
    for p in r["projects"]:
        h = p["health"]
        icon = {"RED": "🔴", "AMBER": "🟡", "GREEN": "🟢"}[h["level"]]
        print(f"  {icon} {p['name']:<20} | {h['label']} | 完成 {h['done_pct']:>4}% / 時間 {h['elapsed_pct']:>4}% | 偏差 {h['forecast_drift']:+4}d")
        print(f"     {p['summary']}")
        a = p["activity"]
        print(f"     里程碑 {len(p['milestones'])} | 延遲 {len(a['delayed'])} 風險 {len(a['risk'])} 進行中 {len(a['active'])} 近完成 {len(a['recent'])}")
    print(f"\n⚠ 跨專案裁示事項：{r['stats']['decision_count']} 件")
    for d in r["decisions"][:5]:
        print(f"   [{d['severity']}] {d['project']:<14} | {d['issue']}")